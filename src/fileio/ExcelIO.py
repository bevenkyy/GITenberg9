# -*- coding:utf-8 -*-

import os
import sys

# sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import openpyxl as opxl
import numpy as np


def query_excel_info(file_path):
    #
    wb = opxl.load_workbook(filename = file_path, read_only = True, data_only = True)
    #
    worksheet_names = wb.sheetnames
    wb.close()
    #
    return worksheet_names

def query_field_index(file_path, ws_name, field):
    '''
    '''
    field_array = read_excel_row(file_path, ws_name, row_index = 0)
    index_arr = np.argwhere(field_array == field)
    #
    return index_arr[0,0]


def read_excel(file_path, ws_name = None, has_row_labels = True, has_col_labels = True):
    '''
    '''
    data = []
    #
    wb = opxl.load_workbook(filename = file_path, read_only = True, data_only = True)
    #
    if ws_name == None:
        ws_name_list = wb.sheetnames
        ws = wb[ws_name_list[0]]  #default:read the first worksheet
    else:
         ws = wb[ws_name]
    #
    if has_row_labels and has_col_labels:
        for i, row in enumerate(ws.rows):
            row_data = []
            if i != 0:
                for j, cell in enumerate(row):
                    if j != 0 and cell.value != None:
                        row_data.append(cell.value)
                #
                if len(row_data) > 0:
                    data.append(row_data)
                
    if not has_row_labels and has_col_labels:
        for i, row in enumerate(ws.rows):
            row_data = []
            if i != 0:
                for cell in row:
                    if cell.value != None:
                        row_data.append(cell.value)
                #
                if len(row_data) > 0:
                    data.append(row_data)
    if has_row_labels and not has_col_labels:
        for row in ws.rows:
            row_data = []
            for j, cell in enumerate(row):
                if j != 0 and cell.value != None:
                    row_data.append(cell.value)
            #
            if len(row_data) > 0:
                data.append(row_data)
    if not has_row_labels and not has_col_labels:
        for row in ws.rows:
            row_data = []
            for cell in row:
                if cell.value != None:
                    row_data.append(cell.value)
            #
            if len(row_data) > 0:
                data.append(row_data)
    #
    data_array = np.array(data, dtype = str)
    #
    wb.close()
    #
    return data_array


def read_excel_row(file_path, ws_name =None, row_index = 0):
    '''
    '''
    wb = opxl.load_workbook(filename = file_path, data_only = True)
    #
    if ws_name == None:
        ws_name_list = wb.sheetnames
        ws = wb[ws_name_list[0]]  #default:read the first worksheet
    else:
        ws = wb[ws_name]
    #
    for i, row in enumerate(ws.rows):
        if i == row_index:
            row_data = [cell.value for cell in row if cell.value != None]
    row_array = np.array(row_data, dtype = str)
    return row_array

def read_excel_col(file_path, ws_name=None, col_index=0):
    '''
    '''
    wb = opxl.load_workbook(filename = file_path, data_only = True)
    if ws_name == None:
        ws_name_list = wb.sheetnames
        ws = wb[ws_name_list[0]]  #default:read the first worksheet
    else:
        ws = wb[ws_name]
    #
    for i, col in enumerate(ws.columns):
        if i == col_index:
            col_data = [cell.value for cell in col if cell.value != None]
    col_array = np.array(col_data, dtype = str)
    return col_array


def write_excel(file_path, ws_name_set, data_set, row_title = None, col_title = None):
    '''
    '''
    #
    wb = opxl.Workbook(write_only = True)
    #
    for ws_name, data in zip(ws_name_set, data_set):
        #
        ws = wb.create_sheet(ws_name)
        #
        if row_title is not None and col_title is not None:
            ws.append(col_title)
            for i in range(data.shape[0]):
                tmp_row_data = data[i,:][np.newaxis,:]
                row_data = tmp_row_data[0].tolist()
                row_data.insert(0, row_title[i])  ##
                ws.append(row_data)
        if row_title is None and col_title is not None:
            ws.append(col_title)
            for i in range(data.shape[0]):
                tmp_row_data = data[i,:][np.newaxis,:]
                row_data = tmp_row_data[0].tolist()
                ws.append(row_data)
        if row_title is not None and col_title is None:
            for i in range(data.shape[0]):
                tmp_row_data = data[i,:][np.newaxis,:]
                row_data = tmp_row_data[0].tolist()
                row_data.insert(0, row_title[i])  ##
                ws.append(row_data)
        if row_title is None and col_title is None:
            for i in range(data.shape[0]):
                tmp_row_data = data[i,:][np.newaxis,:]
                row_data = tmp_row_data[0].tolist()
                ws.append(row_data)
    #
    wb.save(file_path)

if __name__ == "__main__":
    #
    # raise Exception(__file__ + " " + "The script can not be excuted alone!")
    data = read_excel(r"H:\About_硕士研究生工作\论文相关\小论文1_水质参数遥感反演中机器学习模型的特征选择\Data\样本数据\Landsat8\CDOM_Landsat8_Sample\NewFeatures\training-cv_set_band2-5.xlsx")
